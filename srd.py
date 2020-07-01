from openpyxl import Workbook
import openpyxl
import xlrd
import xlwt
import datetime
import cv2
import pytesseract
import re




img = cv2.imread('chech.jpg')
text = pytesseract.image_to_string(img, lang="rus")
f = open('ch.txt','w')
f.write(text)
f.close()
read_photo = text

txt_number = len(open('ch.txt', 'r').readlines())
division = (txt_number//2)

while True:
	index_number = 0
	index = index_number + 1
	print (index)
	regular_index = re.findall(r"[=]\d.\d", read_photo)
	print(regular_index[index_number])
	if index_number == division:
		print('End')
		beak

#разбивка текста 



#открытие файла
dest_filename = 'empty_book.xlsx'
wb=openpyxl.load_workbook(dest_filename)
sheet=wb['Pi']

#запись инфы в первую ячейку Date
row_number = 1 
while True:
	row_number += 1
	column_value = sheet.cell(row=row_number, column=1).value
	if column_value == None:
		print("Свободная первая ячейка --", row_number)
		break

date_info = datetime.datetime.now()
row_update = str(row_number)
sheet['A'+row_update] = date_info


#запись инфы в вторую ячейку Product name
row_number = 1 
while True:
	row_number += 1
	column_value = sheet.cell(row=row_number, column=2).value
	if column_value == None:
		print("Свободная вторая ячейка --", row_number)
		break

product_name = ('Milk')
sheet['B'+row_update] = product_name


#запись инфы в третью ячейку Price
row_number = 1 
while True:
	row_number += 1
	column_value = sheet.cell(row=row_number, column=2).value
	if column_value == None:
		print("Свободная третья ячейка --", row_number)
		break

price = ('3$')
sheet['C'+row_update] = price


wb.save(filename = dest_filename) 